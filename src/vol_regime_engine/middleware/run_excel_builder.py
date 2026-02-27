from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class RunExcelBuilder:

    # --------------------------------------------------
    # Main Builder
    # --------------------------------------------------

    def build(self, run_payload, output_path):

        states = run_payload["states"]
        systemic = run_payload.get("systemic_metrics", {})

        wb = Workbook()

        # ==================================================
        # 1️⃣ RANKING SHEET
        # ==================================================

        ws_rank = wb.active
        ws_rank.title = "Ranking"

        headers = [
            "Underlying",
            "Gamma Regime",
            "Regime Confidence",
            "Acceleration Probability",
            "Signal Strength Index",
            "Regime Persistence Score",
            "Instability Intensity Score",
            "Quadrant Classification"
        ]

        for col, header in enumerate(headers, start=1):
            ws_rank.cell(row=1, column=col, value=header).font = Font(bold=True)

        ranking_rows = []

        for name, state in states.items():

            confidence = state.get("regime_confidence", 0)
            accel = state.get("acceleration_probability", 0)

            ssi = round(0.6 * accel + 0.4 * confidence, 3)
            persistence = round(confidence * (1 - 0.5 * accel), 3)
            instability = round(accel * 100, 1)

            if ssi >= 0.6 and persistence >= 0.6:
                quadrant = "Stable High-Conviction Trend"
            elif ssi >= 0.6:
                quadrant = "Explosive Breakout Candidate"
            elif persistence >= 0.6:
                quadrant = "Range / Mean-Revert Stable"
            else:
                quadrant = "Unstable / Low Edge"

            ranking_rows.append([
                name,
                state.get("gamma_surface_regime"),
                confidence,
                accel,
                ssi,
                persistence,
                instability,
                quadrant
            ])

        # Sort by Signal Strength Index (descending)
        ranking_rows.sort(key=lambda x: x[4], reverse=True)

        for row_idx, row in enumerate(ranking_rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                ws_rank.cell(row=row_idx, column=col_idx, value=val)

        self._auto_adjust_columns(ws_rank)

        # ==================================================
        # 2️⃣ SYSTEMIC DIAGNOSTICS SHEET
        # ==================================================

        ws_sys = wb.create_sheet("Systemic_Diagnostics")

        ws_sys.append(["Metric", "Score"])
        ws_sys["A1"].font = Font(bold=True)
        ws_sys["B1"].font = Font(bold=True)

        for k, v in systemic.items():
            ws_sys.append([k, v])

        self._auto_adjust_columns(ws_sys)

        # ==================================================
        # 3️⃣ INSTABILITY HEATMAP SHEET
        # ==================================================

        ws_heat = wb.create_sheet("Instability_Heatmap")

        ws_heat.append(["Underlying", "Instability Score (0-100)"])
        ws_heat["A1"].font = Font(bold=True)
        ws_heat["B1"].font = Font(bold=True)

        for name, state in states.items():
            accel = state.get("acceleration_probability", 0)
            instability = round(accel * 100, 1)
            ws_heat.append([name, instability])

        self._auto_adjust_columns(ws_heat)

        # ==================================================
        # 4️⃣ DETAILED STATES SHEET
        # ==================================================

        # ws_detail = wb.create_sheet("Detailed_States")
        #
        # ws_detail.append(["Underlying", "Metric", "Value"])
        # ws_detail["A1"].font = Font(bold=True)
        # ws_detail["B1"].font = Font(bold=True)
        # ws_detail["C1"].font = Font(bold=True)
        #
        # for name, state in states.items():
        #     for k, v in state.items():
        #         ws_detail.append([name, k, str(v)])
        #
        # self._auto_adjust_columns(ws_detail)

        # ==================================================
        # 5️⃣ CROSS-ASSET REGIME HEATMAP
        # ==================================================

        from openpyxl.styles import PatternFill

        ws_heatmap = wb.create_sheet("Cross_Asset_Heatmap")

        headers = [
            "Underlying",
            "Gamma Regime",
            "Acceleration",
            "Persistence",
            "Instability",
            "Flip Risk",
            "Early Crash Warning"
        ]

        ws_heatmap.append(headers)

        for col in range(1, len(headers) + 1):
            ws_heatmap.cell(row=1, column=col).font = Font(bold=True)

        systemic_flip = systemic.get("cross_asset_flip_risk", 0)
        systemic_ecws = systemic.get("early_crash_warning", 0)

        row_idx = 2

        for name, state in states.items():
            confidence = state.get("regime_confidence", 0)
            accel = state.get("acceleration_probability", 0)

            persistence = round(confidence * (1 - 0.5 * accel), 3)
            instability = round(accel * 100, 1)

            ws_heatmap.append([
                name,
                state.get("gamma_surface_regime"),
                accel,
                persistence,
                instability,
                systemic_flip,
                systemic_ecws
            ])

            row_idx += 1

        # --------------------------------------------------
        # Apply Color Scaling
        # --------------------------------------------------

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        for row in ws_heatmap.iter_rows(min_row=2, max_row=ws_heatmap.max_row):

            accel_cell = row[2]
            persistence_cell = row[3]
            instability_cell = row[4]
            ecws_cell = row[6]

            # Acceleration coloring
            if accel_cell.value > 0.7:
                accel_cell.fill = red
            elif accel_cell.value > 0.4:
                accel_cell.fill = yellow
            else:
                accel_cell.fill = green

            # Persistence coloring (inverse logic)
            if persistence_cell.value < 0.4:
                persistence_cell.fill = red
            elif persistence_cell.value < 0.6:
                persistence_cell.fill = yellow
            else:
                persistence_cell.fill = green

            # Instability coloring
            if instability_cell.value > 70:
                instability_cell.fill = red
            elif instability_cell.value > 40:
                instability_cell.fill = yellow
            else:
                instability_cell.fill = green

            # Early Crash Warning coloring
            if ecws_cell.value > 0.75:
                ecws_cell.fill = red
            elif ecws_cell.value > 0.5:
                ecws_cell.fill = yellow
            else:
                ecws_cell.fill = green

        self._auto_adjust_columns(ws_heatmap)

        # ==================================================
        # Save
        # ==================================================

        wb.save(str(output_path))
        return output_path

    # --------------------------------------------------
    # Column Auto Width
    # --------------------------------------------------

    def _auto_adjust_columns(self, worksheet):

        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = min(length + 2, 40)